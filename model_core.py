"""
model_core.py — 黑山土石方测算模型的「计算核心」

职责：
    compute(params: dict) -> dict
        输入 21 个用户参数 → 输出单价、各成本拆解、警告等

    export_xlsx(params: dict) -> bytes
        输入 21 个用户参数 → 返回完整 xlsx 文件的字节流（可直接下载）

依赖：
    - build_v8.py（同目录上一级）
    - libreoffice headless（系统命令，用于公式重算）

启动开销：
    首次 import 时跑一次 build_v8 生成 template_v8.xlsx，约 1-2 秒
    后续每次 compute / export 调用约 5-10 秒（libreoffice 重算）
"""

from __future__ import annotations

import os
import sys
import shutil
import tempfile
import subprocess
import io
from typing import Any, Dict

# ---------- 路径：让本模块可定位到 build_v8.py（同目录或上一级目录都能找到）----------
_HERE = os.path.dirname(os.path.abspath(__file__))
_PARENT = os.path.dirname(_HERE)
for _p in (_HERE, _PARENT):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- 启动时生成一次模板 ----------
_TEMPLATE_PATH = os.path.join(_HERE, 'template_v8.xlsx')

def _ensure_template():
    """首次调用时跑 build_v8 生成 template_v8.xlsx；后续直接复用"""
    if os.path.exists(_TEMPLATE_PATH):
        return
    import build_v8  # 执行 build_v8.py 模块（已在 if __name__='__main__' 内保护 save）
    build_v8.wb.save(_TEMPLATE_PATH)

# ---------- 输入参数到主表 cell 的映射 ----------
# (cell, 类型 'str'|'num', 默认值)
PARAM_MAP = {
    # 岩石
    'rock_level':   ('B4',  'str', 'Ⅰ极软岩'),
    # 工艺
    'process':      ('B12', 'str', '直接开挖'),
    # 挖装设备
    'excavator':    ('B17', 'str', 'W4.5'),
    'excavator_src':('B18', 'str', '自有'),
    # 运输设备
    'truck':        ('B23', 'str', 'T65B'),
    'truck_src':    ('B24', 'str', '自有'),
    # 钻孔设备
    'drill':        ('B29', 'str', 'ZG90'),
    'drill_src':    ('B30', 'str', '自有'),
    # 运输参数
    'dist_in':      ('B34', 'num', 2.1),
    'dist_out':     ('B35', 'num', 5.0),
    'dir_in':       ('B37', 'str', '平路'),
    'slope_in':     ('B38', 'num', 1.5),
    'dir_out':      ('B39', 'str', '平路'),
    'slope_out':    ('B40', 'num', 1.0),
    # 爆破设计
    'blast_len':    ('B44', 'num', 50),
    'blast_wid':    ('B45', 'num', 20),
    'buffer_rows':  ('B46', 'num', 1),
    'step_h':       ('B47', 'num', 10),
    'slope_angle':  ('B48', 'num', 75),
    'hole_angle':   ('B49', 'num', 90),
    'hole_pattern': ('B50', 'str', '矩形'),
    'hole_diameter':('B52', 'num', 90),
    'pre_split':    ('B53', 'str', '否'),
}

# ---------- 输出 cell（用于 compute 返回值） ----------
OUTPUT_CELLS = {
    # 综合单价（最关键）
    'price_incl_tax':   ('成本汇总-测算主表', 'B65'),
    'price_excl_tax':   ('成本汇总-测算主表', 'B64'),
    # 各成本项
    'cost_blast':       ('成本汇总-测算主表', 'B57'),
    'cost_excavate':    ('成本汇总-测算主表', 'B58'),
    'cost_transport':   ('成本汇总-测算主表', 'B59'),
    'cost_loosen':      ('成本汇总-测算主表', 'B60'),
    'cost_crush':       ('成本汇总-测算主表', 'B61'),
    'cost_second_crush':('成本汇总-测算主表', 'B62'),
    'cost_dump':        ('成本汇总-测算主表', 'B63'),
    # 联动的岩石参数
    'f_value':          ('成本汇总-测算主表', 'B5'),
    'density':          ('成本汇总-测算主表', 'B6'),
    'loose_factor':     ('成本汇总-测算主表', 'B7'),
    'big_block_rate':   ('成本汇总-测算主表', 'B8'),
    # 推荐
    'recommend_process':('成本汇总-测算主表', 'B13'),
    'recommend_exc':    ('成本汇总-测算主表', 'B19'),
    'recommend_truck':  ('成本汇总-测算主表', 'B22'),
    'recommend_drill':  ('成本汇总-测算主表', 'B28'),
    'recommend_hole_d': ('成本汇总-测算主表', 'B51'),   # 推荐孔径（V8 B51 公式按岩石+工艺推导）
    # 校验警告
    'warn_process':     ('成本汇总-测算主表', 'B14'),
    'warn_truck':       ('成本汇总-测算主表', 'B25'),
    'warn_drill':       ('成本汇总-测算主表', 'B31'),
    'warn_slope':       ('成本汇总-测算主表', 'B41'),
    'warn_step':        ('成本汇总-测算主表', 'B54'),
}

# ---------- 选项常量（给 UI 用） ----------
OPTIONS = {
    'rock_level':   ['Ⅰ极软岩', 'Ⅱ软岩', 'Ⅲ较软岩', 'Ⅳ中硬岩',
                     'Ⅴ较硬岩', 'Ⅵ坚硬岩', 'Ⅶ极硬岩', 'Ⅷ特殊坚硬岩'],
    'process':      ['直接开挖', '松土器松土+开挖', '机械破碎+开挖',
                     '爆破+开挖', '爆破+二次破碎+开挖'],
    'excavator':    ['W1.6', 'W2.0', 'W2.5', 'W3.2', 'E2.0',
                     'W4.5', 'W5.5', 'W6.5', 'W8.0', 'E6.5', 'E8.0'],
    'excavator_src':['自有', '租赁'],
    'truck':        ['T30H', 'E30S', 'T65B', 'T60S', 'T91B', 'E60Y', 'E93L', 'E75T'],
    'truck_src':    ['自有', '租赁'],
    'drill':        ['ZG90', 'ZG120', 'ACD7', 'ACD9', 'AC90', 'AC120', 'KY200', 'KY250'],
    'drill_src':    ['自有', '租赁'],
    'dir_in':       ['平路', '上坡', '下坡'],
    'dir_out':      ['平路', '上坡', '下坡'],
    'hole_pattern': ['矩形', '梅花形'],
    'hole_diameter':[70, 90, 115, 150, 200, 250],
    'pre_split':    ['是', '否'],
    'buffer_rows':  [0, 1, 2],
}

# ---------- 默认参数（用户视角的"默认工况"） ----------
DEFAULT_PARAMS = {k: v[2] for k, v in PARAM_MAP.items()}

# ---------- v10 修复版 CHANGELOG（每次升级追加，便于 Excel 端永久追溯） ----------
CHANGELOG_V10 = [
    {
        'no': 1,
        'date': '2026-06-03',
        'title': '硬岩+松土阶梯定价被绕过',
        'severity': '🔴 严重',
        'files': 'model_core.py (apply_penalty + _inject_correction_layer)',
        'before': 'Ⅴ/Ⅵ/Ⅶ岩+松土时 V8 原算 763 / 1438 / 8912 元/方，远超物理边界',
        'after':  '强制走阶梯定价 baseline×2^gap，Ⅴ=134.92 / Ⅵ=154.36 / Ⅶ=181.08 元/方',
        'detail': '删除 if extra>=0/else 分支，无论 V8 原算多少都按阶梯定价；V8 原算保留为"工艺不可行性证据"显示在校正区',
    },
    {
        'no': 2,
        'date': '2026-06-03',
        'title': '坡度参数失灵（OAT 扫描 0 响应）',
        'severity': '🟡 中等',
        'files': 'template_v8.xlsx 运输 B59/B61 + 参数库 Z25/Z26',
        'before': 'slope 0~25% 全程 B65=11.51（坡度完全不进公式）',
        'after':  '上坡油耗 +B26×Z25(6%/%)，下坡 MAX(0.7, 1-B26×Z26(3%/%))，平路按方向取均值×0.5',
        'detail': 'B59/B61 末尾乘 IF(方向, 上坡修正, 下坡修正, 平路修正) 三分支',
    },
    {
        'no': 3,
        'date': '2026-06-03',
        'title': '爆破几何 3 参数失灵',
        'severity': '🟡 中等',
        'files': 'template_v8.xlsx 钻爆 B17/B52 + 参数库 Z27~Z31',
        'before': '缓冲孔/坡面角/台阶高度三参数 0 响应（仅预裂=是时才用 buffer_rows）',
        'after':  '主爆破费 B17 末尾 ×(1+B22×Z27)；钻孔单耗 B52 乘坡面角和台阶高度修正',
        'detail': '缓冲孔基线 15%/排，坡面角基准 75°、单耗系数 0.008/°，台阶基准 10m、系数 0.015',
    },
    {
        'no': 4,
        'date': '2026-06-03',
        'title': '自有 vs 租赁倒挂',
        'severity': '🟠 反直觉',
        'files': 'template_v8.xlsx 挖装 B34 / 运输 B73 / 钻爆 B98 + 参数库租赁单价',
        'before': '矿卡租赁比自有便宜 1%，钻机租赁便宜 4%（与市场常识相反）',
        'after':  '矿卡租赁比自有贵 1.3%，三处全部 MAX(参数库租赁, 自有×1.05) 兜底',
        'detail': '双保险：①06-02 参数库租赁单价整体上调 ②公式兜底租赁不得低于自有×1.05',
    },
    {
        'no': 5,
        'date': '2026-06-03',
        'title': '诊断脚本"工艺过剩"误报',
        'severity': '⚪ 工具',
        'files': '通测_20260603/脚本/diagnose_v2.py',
        'before': 'user_level>rec_level 时也报红（过剩工艺更贵符合物理却被误判 bug）',
        'after':  '只有 user_level<rec_level 且更便宜才报红；过剩工艺降为🟡观察项',
        'detail': '工艺过剩时单价正常应该更贵，是用户主动选择代价不是 bug，降级观察',
    },
    {
        'no': 6,
        'date': '2026-06-02',
        'title': '参数库锚定真实市场价（康定+宇通试用数据）',
        'severity': '🟢 数据',
        'files': 'template_v8.xlsx 参数库 J26:J33/M26:M33/K40:K47/N40:N47/K51:K58/N51:N58',
        'before': 'W4.5 挖机 320 万 / T65B 矿卡 110 万 / ZG90 钻机 130 万（偏离市场）',
        'after':  'W4.5=250 万 / T65B=80 万 / 50 吨电池矿卡（徐工 385）=96 万',
        'detail': '锚点来源：①徐工康定项目正式报价单 ②宇通 5 月试用车采购数据 ③主流矿卡品牌网询价',
    },
    {
        'no': 7,
        'date': '2026-06-03',
        'title': '孔径↔钻机视觉强化（方案A）',
        'severity': '⚪ 体验',
        'files': 'app.py 孔径 selectbox 后追加 caption + warning + error',
        'before': '换钻机时孔径在悄悄跳，用户察觉不到；爆破区独立选孔径时与钻机脱节，只能事后到「不匹配诊断」里看',
        'after':  '孔径下方实时显示「钻机xxx 推荐 N mm·可用 [...]」，非标配弹黄色 warning，超区间弹红色 error',
        'detail': '不限制孔径下拉选项（保留大马拉小车/小马拉大车对比能力），三档反馈：✅标配 / ⚠️非标配但在区间 / 🚫超区间',
    },
]


# ============================================================
# 设备元信息映射（方案 I 新增 · 2026-06-01）
# ------------------------------------------------------------
# 用于 UI 在选型控件下方展示设备的真实物理含义，避免用户被"代号"误导。
# 比如 E60Y 看着像"更大型号"，但其实是 60t 纯电；E93L 才是 93t 纯电。
# 同时为「钻机↔孔径」联动提供权威映射（来源：build_v8.py 参数库 A40:O58）
# ============================================================

# 矿卡：代号 → (载重t, 动力类型)  ——  数据来源：参数库 A40:E47
TRUCK_INFO = {
    'T30H': (30, '柴油'),
    'E30S': (31, '纯电'),
    'T65B': (65, '柴油'),
    'T60S': (60, '柴油'),
    'T91B': (91, '柴油'),
    'E60Y': (60, '纯电'),
    'E93L': (93, '纯电'),
    'E75T': (75, '纯电'),
}

# 钻机：代号 → (类型, 推荐孔径mm, 孔径区间mm)
#   推荐孔径：默认匹配孔径（用于联动跳变的默认值）
#   孔径区间：参数库 A63:H84 中该钻机覆盖的所有孔径
#   注：潜孔钻 ZG/AC 系列推荐 90mm（最经济）；牙轮钻 KY 系列推荐自身命名口径
DRILL_INFO = {
    'ZG90':  ('潜孔钻', 90,  [70, 90, 115]),
    'ZG120': ('潜孔钻', 90,  [70, 90, 115]),
    'ACD7':  ('潜孔钻', 90,  [70, 90, 115]),
    'ACD9':  ('潜孔钻', 90,  [70, 90, 115]),
    'AC90':  ('潜孔钻', 90,  [70, 90, 115]),
    'AC120': ('潜孔钻', 115, [70, 90, 115]),
    'KY200': ('牙轮钻', 200, [150, 200]),
    'KY250': ('牙轮钻', 250, [200, 250]),
}


def drill_recommend_diameter(drill: str) -> int:
    """根据钻机型号，返回该钻机的推荐孔径（用于 UI 钻机↔孔径联动）"""
    info = DRILL_INFO.get(drill)
    return info[1] if info else 90


def truck_display(code: str) -> str:
    """矿卡代号 → '代号(载重t·动力)' 的展示字符串，UI 用"""
    info = TRUCK_INFO.get(code)
    if not info:
        return code
    return f"{code}（{info[0]}t·{info[1]}）"


def drill_display(code: str) -> str:
    """钻机代号 → '代号(类型·推荐孔径)' 的展示字符串，UI 用"""
    info = DRILL_INFO.get(code)
    if not info:
        return code
    return f"{code}（{info[0]}·标配 {info[1]}mm）"


# ============================================================
# 工艺错配惩罚体系（V8 模型外置·后置校正层）
# ------------------------------------------------------------
# V8 模型按用户选择直接计算，不做合理性校验；这一层在 compute()
# 末尾做两件事：
#   1) 根据"岩石—推荐工艺"映射识别错配
#   2) 对挖装费应用 (1/eff + 0.3*(con-1)) 复合惩罚，并连带刷新总价
#   3) 输出 warning 字段，前端按 severity 决定红/橙/蓝色提示
# ============================================================

# 岩石级别 → 推荐工艺（**完全对齐 V8 参数库 I3:I10**）
# 设计原则（方案H 2026-06-01）：
#   - 工艺类型选择 ≠ 工艺参数设计
#   - V8 已经按岩石硬度自动调整参数（装药量、孔距、钻孔速度等），成本天然不同
#   - 错配校验只针对"工艺类型大方向不对"（用爆破打软岩 / 直接开挖打硬岩）
#   - 大块率高是"爆破设计参数问题"，由工程师调炮孔布置，不属于工艺类型错配
#   - 因此硬岩（含 Ⅷ岩）推荐"爆破+开挖"即合理，二次破碎不是必须
ROCK_RECOMMEND = {
    'Ⅰ极软岩':     '直接开挖',
    'Ⅱ软岩':       '松土器松土+开挖',
    'Ⅲ较软岩':     '松土器松土+开挖',
    'Ⅳ中硬岩':     '爆破+开挖',
    'Ⅴ较硬岩':     '爆破+开挖',
    'Ⅵ坚硬岩':     '爆破+开挖',
    'Ⅶ极硬岩':     '爆破+开挖',
    'Ⅷ特殊坚硬岩':  '爆破+开挖',
}

# 工艺强度档位（数字越小越偷懒）
PROCESS_LEVEL = {
    '直接开挖':           1,
    '松土器松土+开挖':    2,
    '机械破碎+开挖':      3,
    '爆破+开挖':          4,
    '爆破+二次破碎+开挖': 5,
}
# 方案H：错配补偿落到"错配工艺主项"上
# —— 哪个工艺被错配选用，错配补偿就加到该工艺对应的成本项
# 同时维护中文名映射供 UI 显示
COST_FIELD_BY_PROCESS = {
    '直接开挖':           'cost_excavate',     # 挖机硬挖磨耗
    '松土器松土+开挖':    'cost_loosen',       # 松土器磨耗剧增
    '机械破碎+开挖':      'cost_crush',        # 破碎锤效率折损
    '爆破+开挖':          'cost_blast',        # 爆破效果不足
    '爆破+二次破碎+开挖': 'cost_second_crush', # 二次破碎不够
}
# 成本项中文名（供 UI 提示文案使用）
COST_FIELD_LABEL = {
    'cost_excavate':     '挖装费',
    'cost_loosen':       '松土费',
    'cost_crush':        '破碎费',
    'cost_blast':        '爆破费',
    'cost_second_crush': '二次破碎费',
    'cost_transport':    '运输费',
    'cost_dump':         '渣场费',
}
# 方案H：撤销了 SPECIAL_PENALTY_FIELD 特殊映射
# 错配补偿统一按 COST_FIELD_BY_PROCESS 落到"用户选的错配工艺"对应成本项
# 不再做"物理影响"层面的猜测重定向（挖装是最末端环节，不应替别的工艺背锅）

# 推荐工艺基准价（含税，元/方）—— 默认参数下 V8 算出的推荐工艺价
# 用作工艺错配阶梯定价的锚点：错配价 = baseline × (1 + gap × 0.5)
RECOMMEND_BASELINE_INCL = {
    'Ⅰ极软岩':      8.92,
    'Ⅱ软岩':       12.80,
    'Ⅲ较软岩':     15.89,
    'Ⅳ中硬岩':     30.01,
    'Ⅴ较硬岩':     33.73,
    'Ⅵ坚硬岩':     38.59,
    'Ⅶ极硬岩':     45.27,
    'Ⅷ特殊坚硬岩': 60.33,
}

# 工艺错配阶梯定价系数：factor = 2 ** gap
#   差 0 档（推荐）→ ×1
#   差 1 档 → ×2
#   差 2 档 → ×4
#   差 3 档 → ×8
#   差 4 档 → ×16
def _gap_factor(gap):
    return 2 ** gap if gap > 0 else 1.0

# (用户选择, 系统推荐) → (效率折减, 耗材倍率, 严重度, 说明)
# severity: 'error'=红 / 'warning'=橙 / 'info'=蓝
PENALTY = {
    # ↓ "工艺不足"型偏离：用户选了更轻量的工艺，挖机硬挖
    ('直接开挖',         '松土器松土+开挖'):    (0.8,  1.2, 'info',    '岩石较硬，挖机硬挖效率折损约 20%，建议改用松土器'),
    ('直接开挖',         '机械破碎+开挖'):      (0.3,  3.0, 'warning', '岩石需破碎，挖机硬挖效率仅 30%，齿尖磨耗剧增'),
    ('直接开挖',         '爆破+开挖'):         (0.1,  5.0, 'error',   '岩石需爆破，挖机几乎挖不动，齿尖暴损'),
    ('直接开挖',         '爆破+二次破碎+开挖'): (0.08, 6.0, 'error',   '特坚岩需爆破+二次破碎，直接开挖完全不可行'),
    ('松土器松土+开挖',   '机械破碎+开挖'):     (0.5,  2.0, 'warning', '岩石较硬，松土头磨耗剧增，效率仅 50%'),
    ('松土器松土+开挖',   '爆破+开挖'):         (0.2,  3.0, 'error',   '岩石需爆破，松土器极不推荐，效率仅 20%'),
    ('松土器松土+开挖',   '爆破+二次破碎+开挖'): (0.15, 4.0, 'error',  '特坚岩松土完全不可行'),
    ('机械破碎+开挖',     '爆破+开挖'):         (0.5,  2.0, 'warning', '岩石需爆破，机械破碎效率仅 50%'),
    ('机械破碎+开挖',     '爆破+二次破碎+开挖'): (0.4,  2.5, 'warning', '特坚岩机械破碎效率较低且大块多'),
    ('爆破+开挖',         '爆破+二次破碎+开挖'): (0.7,  1.5, 'info',    '特坚岩建议带二次破碎，否则大块多→运输/破碎成本上升'),
    # ↓ "工艺过剩"型偏离：杀鸡牛刀，不影响真实成本，仅提示
    ('松土器松土+开挖',   '直接开挖'):          (1.0,  1.0, 'info', '工艺略偏保守，可改用直接开挖更经济'),
    ('机械破碎+开挖',     '直接开挖'):          (1.0,  1.0, 'info', '工艺过剩，可改用直接开挖更经济'),
    ('机械破碎+开挖',     '松土器松土+开挖'):    (1.0,  1.0, 'info', '工艺偏保守'),
    ('爆破+开挖',         '直接开挖'):          (1.0,  1.0, 'info', '工艺过剩，可改用直接开挖大幅降本'),
    ('爆破+开挖',         '松土器松土+开挖'):    (1.0,  1.0, 'info', '工艺偏保守'),
    ('爆破+开挖',         '机械破碎+开挖'):      (1.0,  1.0, 'info', '工艺偏保守'),
    ('爆破+二次破碎+开挖', '直接开挖'):          (1.0,  1.0, 'info', '工艺过剩'),
    ('爆破+二次破碎+开挖', '松土器松土+开挖'):    (1.0,  1.0, 'info', '工艺过剩'),
    ('爆破+二次破碎+开挖', '机械破碎+开挖'):      (1.0,  1.0, 'info', '工艺偏保守'),
    ('爆破+二次破碎+开挖', '爆破+开挖'):          (1.0,  1.0, 'info', '工艺偏保守'),
}


def detect_mismatch(rock_level, process):
    """识别工艺错配。返回 dict 或 None。"""
    recommended = ROCK_RECOMMEND.get(rock_level)
    if not recommended or process == recommended:
        return None
    rule = PENALTY.get((process, recommended))
    if not rule:
        return None
    eff, con, sev, msg = rule
    return {
        'severity': sev,
        'message': msg,
        'process_user': process,
        'process_recommend': recommended,
        'rock_level': rock_level,
        'efficiency_factor': eff,
        'consume_factor': con,
        'is_penalty': (eff < 1.0 or con > 1.0),
    }


def apply_penalty(result):
    """工艺错配后置校正：按"工艺档位差"覆盖式定价。

    业务原则：工艺越偷懒（强度越低），单价应该越贵。
        直接开挖 > 松土器 > 机械破碎 > 爆破+开挖

    实现：
    - 每个岩石等级有「推荐工艺」和「推荐工艺基准价」
    - 用户选其它工艺时，按档位差强制定价：基准价 × (1 + gap × 0.5)
    - 覆盖 V8 原算，避免「V8 在硬岩+松土器自带产量罚导致单价飙到几百几千」
      和「V8 在硬岩+直接开挖完全没罚导致直接开挖比松土器还便宜」两个 bug
    - V8 原始价记录在 warning.v8_original_price_incl，让用户看到「原始算 763 元/方」的不可行性
    """
    inputs = result.get('inputs', {})
    rock = inputs.get('rock_level')
    proc = inputs.get('process')
    mismatch = detect_mismatch(rock, proc)
    result['warning'] = mismatch
    if not mismatch:
        return result

    level_user = PROCESS_LEVEL.get(proc, 0)
    level_rec = PROCESS_LEVEL.get(mismatch['process_recommend'], 0)
    gap = level_rec - level_user  # 正数=用户工艺偏软（不足）

    base_excl = result.get('price_excl_tax') or 0
    base_incl = result.get('price_incl_tax') or 0
    tax_ratio = (base_incl / base_excl) if base_excl > 0 else 1.09

    # ---- V8 模型跳过项检测（用于 UI 解释「为什么选了爆破但没爆破费」）----
    # V8 模型 B57 公式：IF(f<2 且选爆破工艺, 0, 计算爆破费) —— f<2 时强制爆破费=0
    # V8 模型 B62 公式：爆破+二次破碎工艺时 = 破碎费×大块率，软岩大块率≈0
    f_value = result.get('f_value') or 0
    skipped = []
    if f_value > 0 and f_value < 2:
        if proc in ('爆破+开挖', '爆破+二次破碎+开挖') and (result.get('cost_blast') or 0) == 0:
            skipped.append('爆破费')
        if proc == '爆破+二次破碎+开挖' and (result.get('cost_second_crush') or 0) == 0:
            skipped.append('二次破碎费')
    if skipped:
        mismatch['v8_skipped_costs'] = skipped
        mismatch['v8_skip_reason'] = (
            f"模型规则：当前岩石普氏系数 f={f_value:.2f} < 2（极软岩自带破碎），"
            f"虽然您选择了「{proc}」，但模型不计算 {('、'.join(skipped))}（取 0）"
        )

    # ---- C 类：工艺过剩（杀鸡牛刀）----
    # 校正层不改价，但记录"推荐工艺基准价"供 UI 做对比，解释为什么 V8 算出的价
    # 和推荐工艺仍有差异（V8 按用户选的工艺真实算各项成本）
    if gap <= 0:
        mismatch['severity'] = 'info'
        # 给 C 类补上对比信息：V8 按用户工艺算的实际价 vs 推荐工艺参考价
        recommend_baseline = RECOMMEND_BASELINE_INCL.get(rock)
        if recommend_baseline is not None:
            mismatch['v8_actual_price_incl'] = round(base_incl, 2)
            mismatch['recommend_baseline_incl'] = round(recommend_baseline, 2)
            mismatch['excess_levels'] = -gap  # 正数=用户工艺过剩档位
            mismatch['note'] = (
                f"工艺过剩 {-gap} 档。按您选的「{proc}」算实际成本 "
                f"{base_incl:.2f} 元/方（含税），推荐工艺「{mismatch['process_recommend']}」"
                f"默认参数下成本约 {recommend_baseline:.2f} 元/方，"
                f"两者差异来自不同工艺算的物料参数（松散系数、大块率等）"
            )
        return result

    # ---- B 类：工艺偏软 → 按档位差强制定价 ----
    baseline_incl = RECOMMEND_BASELINE_INCL.get(rock)
    if baseline_incl is None:
        return result  # 安全 fallback
    baseline_excl = baseline_incl / tax_ratio
    factor = _gap_factor(gap)
    target_excl = baseline_excl * factor
    target_incl = target_excl * tax_ratio

    # 记录到 warning 供 UI 显示
    mismatch['v8_original_price_incl'] = round(base_incl, 2)
    mismatch['baseline_price_incl'] = round(baseline_incl, 2)
    mismatch['gap_levels'] = gap
    mismatch['adjusted_factor'] = round(factor, 2)
    mismatch['adjusted_price_incl'] = round(target_incl, 2)

    # 方案H：错配补偿统一落到"用户选的错配工艺主项"上（物理纯净 + 不踩挖装）
    # —— 选错松土器就罚松土费、选错破碎就罚破碎费、选错爆破就罚爆破费
    # —— 不再做"物理影响"猜测重定向（挖装作为最末端环节不替别人背锅）
    # 边界保护：当 V8 原算 ≥ 阶梯定价时按 V8 实际成本走（避免出现负数）
    main_field = COST_FIELD_BY_PROCESS.get(proc, 'cost_excavate')
    main_label = COST_FIELD_LABEL.get(main_field, main_field)
    orig_main = result.get(main_field) or 0
    # ---- v10 修复（病灶1）：强制走阶梯定价 ----
    # 旧逻辑：MAX(target_excl, base_excl) → V8原算更高时不下调
    # 问题：V8在硬岩+松土自带产量罚，原算763/1438/8912，全部走else分支不下调
    # 新逻辑：无论V8原算多少，错配工艺一律按 baseline × 2^gap 强制定价
    #         V8原算保留在 warning.v8_original_price_incl，UI 显示作为"工艺不可行性证据"
    extra = target_excl - base_excl
    result[main_field + '_original'] = orig_main
    result[main_field] = orig_main + extra  # extra 可能为负（V8原算>阶梯定价时）
    result['penalty_field'] = main_field
    result['penalty_field_label'] = main_label
    result['penalty_extra'] = round(extra, 4)
    result['penalty_multiplier'] = factor
    result['price_excl_tax'] = target_excl
    result['price_incl_tax'] = target_incl
    mismatch['penalty_field'] = main_field
    mismatch['penalty_field_label'] = main_label
    mismatch['penalty_extra'] = round(extra, 2)
    mismatch['penalty_field_original'] = round(orig_main, 2)
    mismatch['penalty_field_adjusted'] = round(orig_main + extra, 2)
    if extra >= 0:
        mismatch['note'] = (
            f"错配补偿 {extra:.2f} 元/方 已加到「{main_label}」上"
            f"（原算 {orig_main:.2f} → 校正后 {orig_main+extra:.2f}）"
        )
    else:
        # V8原算 > 阶梯定价：强制下调到阶梯定价，原算作为"工艺不可行性证据"
        mismatch['note'] = (
            f"V8 原算 {base_incl:.2f} 元/方（错配工艺产量极低导致台班费÷低产量飙高），"
            f"已强制按阶梯定价 {target_incl:.2f} 元/方收（baseline {baseline_incl:.2f} × 2^{gap}）。"
            f"V8 原算反向证明该工艺不可行，建议改用「{mismatch['process_recommend']}」"
        )
    return result


# ---------- 核心：把参数应用到模板 + libreoffice 重算 ----------
def _build_and_calc(params: Dict[str, Any], workdir: str) -> str:
    """把参数写入主表 → save → libreoffice 重算 → 返回重算后文件路径"""
    _ensure_template()
    src = os.path.join(workdir, 'src.xlsx')
    shutil.copy(_TEMPLATE_PATH, src)

    # 1. 写主表参数
    wb = load_workbook(src)
    ws = wb['成本汇总-测算主表']
    full = {**DEFAULT_PARAMS, **(params or {})}
    for key, val in full.items():
        if key not in PARAM_MAP:
            continue
        cell, typ, _ = PARAM_MAP[key]
        if typ == 'num':
            try:
                val = float(val)
            except (TypeError, ValueError):
                continue
        ws[cell] = val
    wb.save(src)

    # 2. libreoffice headless 重算
    out_dir = os.path.join(workdir, 'recalc')
    os.makedirs(out_dir, exist_ok=True)
    soffice = _find_libreoffice()
    cmd = [soffice, '--headless', '--calc',
           '--convert-to', 'xlsx', '--outdir', out_dir, src]
    res = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    if res.returncode != 0:
        raise RuntimeError(f"libreoffice 重算失败：{res.stderr or res.stdout}")
    out_path = os.path.join(out_dir, 'src.xlsx')
    if not os.path.exists(out_path):
        raise RuntimeError("libreoffice 未生成输出文件")
    return out_path


def _find_libreoffice() -> str:
    for name in ('soffice', 'libreoffice'):
        p = shutil.which(name)
        if p:
            return p
    raise RuntimeError("找不到 libreoffice/soffice 命令，请先安装 libreoffice")


# ============================================================
# 共享审计报告构造（网页 + Excel 统一调用）
# -----------------------------------------------------------
# 把推荐组合 vs 用户选择的 7 维度审计逻辑集中在此，
# app.py 的 UI 和 export_xlsx 的 Excel sheet 都调用同一个函数，
# 保证"改一处，两端自动同步"。
# ============================================================

def build_audit_report(result: Dict[str, Any]) -> Dict[str, Any]:
    """根据 compute() 返回的结果，构造完整的审计报告。

    返回 dict 结构：
    {
        'mismatches': [ (维度名, 用户选, 推荐, 原因), ... ],
        'process_extra_lines': [ str, ... ],   # 工艺错配补偿明细（紧跟工艺条目）
        'all_match': bool,                      # 是否全部匹配
    }
    """
    warning = result.get('warning') or {}
    inputs = result.get('inputs', {})

    rec_process = result.get('recommend_process', '—')
    rec_exc     = result.get('recommend_exc', '—')
    rec_truck   = result.get('recommend_truck', '—')
    rec_drill   = result.get('recommend_drill', '—')
    rec_hole_d  = result.get('recommend_hole_d', '—')

    user_process = inputs.get('process', '—')
    user_exc     = inputs.get('excavator', '—')
    user_truck   = inputs.get('truck', '—')
    user_drill   = inputs.get('drill', '—')
    user_hole_d  = inputs.get('hole_diameter', '—')

    mismatches = []

    # 1) 工艺
    if user_process != rec_process:
        gap = warning.get('gap_levels') or warning.get('excess_levels', 0)
        phys_hint = ""
        if warning.get('severity') in ('error', 'warning') and gap and gap >= 3:
            phys_hint = "；此组合在物理上几乎不可行（岩石过硬，弱工艺无法奏效），需按推荐工艺的代价补偿"
        if warning.get('severity') in ('error', 'warning'):
            reason = f"当前岩石推荐「{rec_process}」，您选了「{user_process}」（偏软 {gap} 档），成本将大幅上升{phys_hint}"
        elif warning.get('severity') == 'info' and warning.get('excess_levels'):
            reason = f"当前岩石推荐「{rec_process}」，您选了「{user_process}」（过剩 {gap} 档，杀鸡用牛刀）"
        elif warning.get('severity') == 'info':
            reason = f"当前岩石推荐「{rec_process}」，您选了「{user_process}」（偏软 {gap} 档）"
        else:
            reason = f"推荐「{rec_process}」vs 您选「{user_process}」，工艺不匹配"
        mismatches.append(("工艺", user_process, rec_process, reason))

    # 2) 挖机
    if user_exc != rec_exc:
        mismatches.append(("挖机", user_exc, rec_exc,
            f"推荐挖机「{rec_exc}」按工艺+岩石强度匹配；您选「{user_exc}」，可能影响挖装效率"))

    # 3) 矿卡（含动力类型判断）
    user_t = TRUCK_INFO.get(user_truck)
    rec_t_code = rec_truck
    for _code in TRUCK_INFO:
        if _code in str(rec_truck):
            rec_t_code = _code
            break
    rec_t = TRUCK_INFO.get(rec_t_code) if rec_t_code in TRUCK_INFO else None

    if user_truck != rec_t_code:
        reason_parts = []
        if user_t and rec_t:
            u_load, u_power = user_t
            r_load, r_power = rec_t
            if u_load != r_load:
                reason_parts.append(f"载重：您选 {user_truck}({u_load}t) vs 推荐 {rec_t_code}({r_load}t)")
            if u_power != r_power:
                reason_parts.append(
                    f"动力：您选 {u_power} vs 推荐 {r_power}"
                    + ("（纯电单方比柴油便宜约 4-5 元/方，但需充电桩配套）" if r_power == '纯电' else "")
                )
        if not reason_parts:
            reason_parts.append(f"推荐「{rec_truck}」vs 您选「{user_truck}」")
        dist_total = (inputs.get('dist_in', 0) or 0) + (inputs.get('dist_out', 0) or 0)
        if dist_total < 3 and user_t and user_t[0] >= 60:
            reason_parts.append("综合运距 <3km + 载重 ≥60t → 大马拉小车不经济")
        elif dist_total > 10 and user_t and user_t[0] < 60:
            reason_parts.append("综合运距 >10km + 载重 <60t → 频繁往返效率低")
        mismatches.append(("矿卡", user_truck, rec_truck, "；".join(reason_parts)))

    # 4) 钻机
    user_d = DRILL_INFO.get(user_drill)
    if rec_drill not in ('—', '', None) and user_drill != rec_drill:
        d_reason = f"推荐钻机「{rec_drill}」按岩石硬度匹配"
        if user_d:
            rec_d = DRILL_INFO.get(rec_drill)
            if rec_d:
                d_reason += f"（{rec_d[0]}·标配{rec_d[1]}mm）"
            d_reason += f"；您选「{user_drill}」({user_d[0]}·标配{user_d[1]}mm)"
        mismatches.append(("钻机", user_drill, rec_drill, d_reason))

    # 5) 孔径（双重判断：vs 推荐孔径 + vs 钻机标配孔径）
    if user_hole_d not in ('—', '', None) and rec_hole_d not in ('—', '', None):
        hole_mismatch = False
        hole_reasons = []
        try:
            uh = int(user_hole_d)
            rh = int(rec_hole_d)
            if uh != rh:
                hole_mismatch = True
                diff = uh - rh
                if diff < 0:
                    hole_reasons.append(
                        f"孔径偏小：您选 {uh}mm < 推荐 {rh}mm，单孔方量减少→每方炸药消耗增加→爆破费升高"
                    )
                else:
                    hole_reasons.append(
                        f"孔径偏大：您选 {uh}mm > 推荐 {rh}mm，需确认钻机能支撑此孔径，否则大孔径+小钻机=小马拉大车"
                    )
        except (ValueError, TypeError):
            pass
        if user_d:
            d_rec = user_d[1]
            d_range = user_d[2]
            try:
                uh2 = int(user_hole_d)
                if uh2 != d_rec:
                    hole_mismatch = True
                    if uh2 < d_rec:
                        hole_reasons.append(
                            f"大马拉小车：钻机「{user_drill}」标配 {d_rec}mm，"
                            f"您选 {uh2}mm → 台班费高但钻速发挥不出来，反而贵"
                        )
                    elif uh2 > d_rec:
                        if uh2 not in d_range:
                            hole_reasons.append(
                                f"小马拉大车：钻机「{user_drill}」可用孔径 {d_range}mm，"
                                f"您选 {uh2}mm 超出范围，可能无法施工"
                            )
                        else:
                            hole_reasons.append(
                                f"钻机「{user_drill}」标配 {d_rec}mm，"
                                f"您选 {uh2}mm（在可用范围内但非标配，台班效率可能下降）"
                            )
            except (ValueError, TypeError):
                pass
        if hole_mismatch:
            mismatches.append(("孔径", f"{user_hole_d}mm", f"{rec_hole_d}mm（推荐）", "；".join(hole_reasons)))

    # 6) 坡度
    if result.get('warn_slope') and '⚠' in str(result.get('warn_slope', '')):
        mismatches.append(("坡度", "当前设置", "合理范围", str(result['warn_slope'])))

    # 7) 台阶/边坡
    if result.get('warn_step') and '⚠' in str(result.get('warn_step', '')):
        mismatches.append(("台阶", "当前设置", "合理范围", str(result['warn_step'])))

    # —— 工艺错配补偿明细（绑定到"工艺"那条 mismatch 下方）——
    process_extra_lines = []
    if warning:
        _gap = warning.get('gap_levels') or warning.get('excess_levels', 0)
        _factor = warning.get('adjusted_factor')
        _baseline = warning.get('baseline_price_incl')
        _adj = warning.get('adjusted_price_incl')
        _v8_raw = warning.get('v8_original_price_incl')
        _ptype = warning.get('process_user', '—')
        _prec = warning.get('process_recommend', '—')

        # B 类（偏软）阶梯定价说明
        if _gap and _factor and _baseline:
            process_extra_lines.append(
                f"阶梯定价：推荐工艺基准 {_baseline:.2f} × 档位差 {_gap} 档系数 {_factor:.2f} = {_adj:.2f} 元/方"
                + (f" ｜ 按您选工艺直接算（含产量罚）：{_v8_raw:.2f} 元/方" if _v8_raw else "")
            )

        # B 类错配补偿落点
        _p_label = warning.get('penalty_field_label')
        _p_extra = warning.get('penalty_extra')
        _p_orig = warning.get('penalty_field_original')
        _p_adj  = warning.get('penalty_field_adjusted')
        if _p_label and _p_extra is not None and _p_extra > 0:
            if _p_orig is not None and _p_adj is not None:
                process_extra_lines.append(
                    f"因工艺错配补偿：{_p_extra:.2f} 元/方 已加到「{_p_label}」上"
                    f"（原算 {_p_orig:.2f} → 校正后 {_p_adj:.2f}），"
                    f"为反映物理不可行的真实代价（其他成本项保持原算真实值不变）"
                )
            else:
                process_extra_lines.append(
                    f"因工艺错配补偿：{_p_extra:.2f} 元/方 已加到「{_p_label}」上，"
                    f"为反映物理不可行的真实代价（其他成本项保持原算真实值不变）"
                )

        # 模型规则跳过项说明
        _skip_reason = warning.get('v8_skip_reason')
        if _skip_reason:
            process_extra_lines.append(_skip_reason.replace('V8', '模型'))

        # C 类工艺过剩差异明细
        if warning.get('severity') == 'info' and warning.get('excess_levels'):
            _v8_actual = warning.get('v8_actual_price_incl')
            _rec_base = warning.get('recommend_baseline_incl')
            if _v8_actual is not None and _rec_base is not None:
                _diff = _v8_actual - _rec_base
                _sign = '+' if _diff >= 0 else ''
                process_extra_lines.append(
                    f"按您选「{_ptype}」算实际成本 {_v8_actual:.2f} vs "
                    f"推荐「{_prec}」默认参数约 {_rec_base:.2f}，"
                    f"差异 {_sign}{_diff:.2f} 元/方（未额外加罚，显示的是真实成本）"
                )

    return {
        'mismatches': mismatches,
        'process_extra_lines': process_extra_lines,
        'all_match': len(mismatches) == 0,
    }


# ---------- 对外 API ----------
def compute(params: Dict[str, Any] = None) -> Dict[str, Any]:
    """
    给一组用户参数，返回所有关键计算结果。
    返回的 dict 包含 OUTPUT_CELLS 里全部 key + 'inputs' 回显。
    """
    with tempfile.TemporaryDirectory() as workdir:
        recalc_path = _build_and_calc(params or {}, workdir)
        wb = load_workbook(recalc_path, data_only=True)
        out = {}
        for key, (sheet, cell) in OUTPUT_CELLS.items():
            v = wb[sheet][cell].value
            out[key] = v
        out['inputs'] = {**DEFAULT_PARAMS, **(params or {})}
        # 应用工艺错配惩罚（加 warning 字段、修正挖装费和总价）
        apply_penalty(out)
        return out


# -----------------------------------------------------------
# Excel 审计 sheet 写入辅助函数
# 让 Excel 下载和网页 UI 共享同一份 build_audit_report 输出
# -----------------------------------------------------------

_FILL_HEAD   = PatternFill('solid', fgColor='305496')   # 表头蓝
_FILL_WARN   = PatternFill('solid', fgColor='FFF2CC')   # 警告黄
_FILL_OK     = PatternFill('solid', fgColor='E2EFDA')   # 匹配绿
_FILL_EXTRA  = PatternFill('solid', fgColor='FCE4D6')   # 错配补偿橙
_FILL_NOTE   = PatternFill('solid', fgColor='F2F2F2')   # 说明灰
_FONT_HEAD   = Font(bold=True, color='FFFFFF', size=11)
_FONT_TITLE  = Font(bold=True, size=12, color='305496')
_FONT_BOLD   = Font(bold=True, size=11)
_ALIGN_WRAP  = Alignment(wrap_text=True, vertical='top', horizontal='left')
_ALIGN_CTR   = Alignment(horizontal='center', vertical='center')
_BORDER_THIN = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4'),
)

# 维度 emoji（与 app.py UI 保持一致）
_CAT_EMOJI = {
    '工艺': '🔧', '挖机': '⛏️', '矿卡': '🚛',
    '钻机': '🛠️', '孔径': '📏', '坡度': '⛰️', '台阶': '📐',
}


def _write_audit_sheet(wb, audit: Dict[str, Any], result: Dict[str, Any]) -> None:
    """写入「📋 工艺设备匹配度校验」sheet。"""
    sheet_name = '📋 工艺设备匹配度校验'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # 列宽
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 60

    # 标题
    ws['A1'] = '工艺设备匹配度校验'
    ws['A1'].font = _FONT_TITLE
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 22

    # 快照提示
    ws['A2'] = ('⚠️ 本表是您下载时的参数快照；如在 Excel 中修改了岩石/工艺等参数，'
                '主表 B65「综合单价(含税)」会自动重算，但本审计表不会跟着变——'
                '想看新参数的校验，请回网页修改参数后重新下载。')
    ws['A2'].font = Font(italic=True, color='C00000', size=10)
    ws['A2'].alignment = _ALIGN_WRAP
    ws['A2'].fill = _FILL_NOTE
    ws.merge_cells('A2:D2')
    ws.row_dimensions[2].height = 36

    # 表头（A2 是快照提示，表头顺延到第 3 行）
    headers = ['维度', '您的选择', '推荐配置', '原因说明']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = _FILL_HEAD
        c.font = _FONT_HEAD
        c.alignment = _ALIGN_CTR
        c.border = _BORDER_THIN
    ws.row_dimensions[3].height = 22

    mismatches = audit.get('mismatches') or []
    extra_lines = audit.get('process_extra_lines') or []

    row = 4
    if not mismatches:
        ws.cell(row=row, column=1, value='✅ 全部匹配').font = _FONT_BOLD
        ws.cell(row=row, column=1).fill = _FILL_OK
        ws.cell(row=row, column=2, value='').fill = _FILL_OK
        ws.cell(row=row, column=3, value='').fill = _FILL_OK
        ws.cell(row=row, column=4,
                value='您选择的工艺、设备、孔径、坡度、台阶等参数与推荐组合完全一致，无需调整。').fill = _FILL_OK
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = _BORDER_THIN
            ws.cell(row=row, column=col).alignment = _ALIGN_WRAP
        row += 1
    else:
        for cat, user_val, rec_val, reason in mismatches:
            emoji = _CAT_EMOJI.get(cat, '⚠️')
            ws.cell(row=row, column=1, value=f"{emoji} {cat}")
            ws.cell(row=row, column=2, value=str(user_val) if user_val is not None else '—')
            ws.cell(row=row, column=3, value=str(rec_val) if rec_val is not None else '—')
            ws.cell(row=row, column=4, value=str(reason) if reason else '')
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.fill = _FILL_WARN
                cell.border = _BORDER_THIN
                cell.alignment = _ALIGN_WRAP
            ws.cell(row=row, column=1).font = _FONT_BOLD
            row += 1
            # 工艺条目下方紧贴错配补偿明细
            if cat == '工艺' and extra_lines:
                for line in extra_lines:
                    ws.cell(row=row, column=1, value='💰 补偿明细')
                    ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=4)
                    ws.cell(row=row, column=2, value=str(line))
                    for col in range(1, 5):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = _FILL_EXTRA
                        cell.border = _BORDER_THIN
                        cell.alignment = _ALIGN_WRAP
                    ws.cell(row=row, column=1).font = _FONT_BOLD
                    row += 1

    # 判定依据
    row += 1
    ws.cell(row=row, column=1,
            value='判定依据').font = _FONT_BOLD
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
    ws.cell(row=row, column=1).fill = _FILL_NOTE
    row += 1
    notes = [
        '• 推荐组合：根据岩石普氏系数（Ⅰ~Ⅷ级）自动推导，含工艺、挖机、矿卡、钻机、孔径、坡度、台阶；',
        '• 工艺错配：偏软档位会触发"阶梯定价"补偿（×2 / ×4 / ×8 / ×16），物理不可行情况在此特别提示；',
        '• 校正后含税单价已在「成本汇总-测算主表」B65 直接覆盖，所见即所得，无需二次重算；',
        '• 想换工艺/设备组合得到不同单价，请回到网页修改参数后重新生成本 Excel。',
    ]
    for line in notes:
        ws.cell(row=row, column=1, value=line).alignment = _ALIGN_WRAP
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
        ws.cell(row=row, column=1).fill = _FILL_NOTE
        row += 1


def _write_correction_sheet(wb, result: Dict[str, Any], audit: Dict[str, Any]) -> None:
    """写入「💰 校正后定价说明」sheet。"""
    sheet_name = '💰 校正后定价说明'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 50

    warning = result.get('warning') or {}
    inputs = result.get('inputs', {})
    rock = inputs.get('rock_level', '—')
    proc = inputs.get('process', '—')
    rec_proc = result.get('recommend_process', '—')

    ws['A1'] = '校正后定价说明'
    ws['A1'].font = _FONT_TITLE
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 22

    # 快照提示
    ws['A2'] = ('⚠️ 本表是您下载时的参数快照；如在 Excel 中修改了岩石/工艺等参数，'
                '主表 B65 会自动重算，但本说明表不会跟着更新——'
                '校正过程的实时数值请看主表 D64:G70 区。')
    ws['A2'].font = Font(italic=True, color='C00000', size=10)
    ws['A2'].alignment = _ALIGN_WRAP
    ws['A2'].fill = _FILL_NOTE
    ws.merge_cells('A2:D2')
    ws.row_dimensions[2].height = 36

    row = 3
    # 当前定价
    ws.cell(row=row, column=1, value='当前定价').font = _FONT_BOLD
    ws.cell(row=row, column=1).fill = _FILL_HEAD
    ws.cell(row=row, column=1).font = _FONT_HEAD
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
    row += 1
    pairs = [
        ('岩石等级',   rock),
        ('选用工艺',   proc),
        ('推荐工艺',   rec_proc),
        ('含税综合单价（元/方）', f"{result.get('price_incl_tax') or 0:.2f}"),
        ('不含税综合单价（元/方）', f"{result.get('price_excl_tax') or 0:.2f}"),
    ]
    for k, v in pairs:
        ws.cell(row=row, column=1, value=k).font = _FONT_BOLD
        ws.cell(row=row, column=2, value=v)
        ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=4)
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = _BORDER_THIN
            ws.cell(row=row, column=col).alignment = _ALIGN_WRAP
        row += 1

    # 错配补偿说明
    if warning and warning.get('severity') != 'info' and warning.get('adjusted_price_incl') is not None:
        row += 1
        ws.cell(row=row, column=1, value='错配补偿明细').font = _FONT_HEAD
        ws.cell(row=row, column=1).fill = _FILL_HEAD
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
        row += 1
        v8_orig = warning.get('v8_original_price_incl')
        baseline = warning.get('baseline_price_incl')
        gap = warning.get('gap_levels')
        factor = warning.get('adjusted_factor')
        adjusted = warning.get('adjusted_price_incl')
        rows = [
            ('V8 原算含税单价（元/方）', f"{v8_orig:.2f}" if v8_orig is not None else '—',
             '未经校正、按用户工艺直接计算（物理上不可行，仅作参考）'),
            ('推荐工艺基准含税单价（元/方）', f"{baseline:.2f}" if baseline is not None else '—',
             f"推荐工艺「{rec_proc}」在 {rock} 下的基准含税单价"),
            ('档位差', f"偏软 {gap} 档" if gap is not None else '—',
             '工艺强度档位差：直接开挖<松土器<机械破碎<爆破+开挖'),
            ('阶梯系数', f"×{factor:.2f}" if factor is not None else '—',
             '档位差对应的强制定价倍率（偏1档×2 / 偏2档×4 / 偏3档×8 / 偏4档×16）'),
            ('校正后含税单价（元/方）', f"{adjusted:.2f}" if adjusted is not None else '—',
             '基准单价 × 阶梯系数，已覆盖到「成本汇总-测算主表」B65'),
        ]
        for k, v, note in rows:
            ws.cell(row=row, column=1, value=k).font = _FONT_BOLD
            ws.cell(row=row, column=2, value=v)
            ws.cell(row=row, column=3, value='')
            ws.cell(row=row, column=4, value=note)
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.fill = _FILL_EXTRA
                cell.border = _BORDER_THIN
                cell.alignment = _ALIGN_WRAP
            row += 1

        # 错配补偿落点
        pen_label = warning.get('penalty_field_label')
        pen_orig  = warning.get('penalty_field_original')
        pen_adj   = warning.get('penalty_field_adjusted')
        pen_extra = warning.get('penalty_extra')
        if pen_label and pen_orig is not None and pen_adj is not None:
            row += 1
            ws.cell(row=row, column=1, value='错配补偿落点').font = _FONT_BOLD
            ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
            ws.cell(row=row, column=1).fill = _FILL_NOTE
            row += 1
            落点行 = [
                ('补偿挂在哪个成本项', pen_label),
                ('原算值（元/方）', f"{pen_orig:.2f}"),
                ('校正后值（元/方）', f"{pen_adj:.2f}"),
                ('增量（元/方）', f"+{pen_extra:.2f}" if pen_extra is not None else '—'),
            ]
            for k, v in 落点行:
                ws.cell(row=row, column=1, value=k).font = _FONT_BOLD
                ws.cell(row=row, column=2, value=v)
                ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=4)
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = _BORDER_THIN
                    ws.cell(row=row, column=col).alignment = _ALIGN_WRAP
                row += 1

    # C 类工艺过剩说明
    elif warning and warning.get('severity') == 'info' and warning.get('excess_levels'):
        row += 1
        ws.cell(row=row, column=1, value='工艺过剩说明').font = _FONT_HEAD
        ws.cell(row=row, column=1).fill = _FILL_HEAD
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
        row += 1
        note_text = warning.get('note', '')
        ws.cell(row=row, column=1, value=note_text).alignment = _ALIGN_WRAP
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
        ws.cell(row=row, column=1).fill = _FILL_EXTRA
        ws.cell(row=row, column=1).border = _BORDER_THIN
        row += 1

    # 模型跳过提示
    skipped = warning.get('v8_skipped_costs') if warning else None
    if skipped:
        row += 1
        ws.cell(row=row, column=1, value='模型跳过的成本项').font = _FONT_BOLD
        ws.cell(row=row, column=1).fill = _FILL_NOTE
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
        row += 1
        ws.cell(row=row, column=1, value=warning.get('v8_skip_reason', '')).alignment = _ALIGN_WRAP
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
        ws.cell(row=row, column=1).fill = _FILL_NOTE
        row += 1

    # 兜底说明
    row += 1
    notes = [
        '• 主表 B64/B65 已直接显示校正后单价，与网页端「💵 校正后含税综合单价」一致；',
        '• 若想换组合得到不同单价，请回到网页修改参数后重新下载本 Excel。',
    ]
    for line in notes:
        ws.cell(row=row, column=1, value=line).alignment = _ALIGN_WRAP
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
        ws.cell(row=row, column=1).fill = _FILL_NOTE
        row += 1


def _inject_correction_layer(wb) -> None:
    """把校正逻辑注入 Excel：写入隐藏配置 sheet + 改主表 B65 为校正后活公式。

    单一数据源核心：RECOMMEND_BASELINE_INCL / PROCESS_LEVEL 来自 model_core.py，
    每次生成 Excel 都会写最新值到「⚙️校正配置」sheet；
    主表 B65 公式 VLOOKUP 这张配置 sheet，所以 model_core 改了基准价/工艺等级，
    用户拿到的 Excel 在本地改参数算出来的结果也会跟着变。

    主表改动（保留所有 V8 原算公式作为基础）：
        B65 公式从 "=B64*(1+税率)"
        改为 "=IF(校正条件, MAX(基准价×系数, V8原算含税), V8原算含税)"

    在主表 D66:E68 区追加可见的"校正过程说明"（供用户理解阶梯定价是怎么算出来的）。
    """
    main_sheet_name = '成本汇总-测算主表'
    cfg_sheet_name = '⚙️校正配置'

    # ① 隐藏配置 sheet（如已存在先删）
    if cfg_sheet_name in wb.sheetnames:
        del wb[cfg_sheet_name]
    cfg = wb.create_sheet(cfg_sheet_name)
    cfg.sheet_state = 'hidden'

    # 标题
    cfg['A1'] = '岩石级别'
    cfg['B1'] = '推荐工艺基准含税价(元/方)'
    cfg['D1'] = '工艺名称'
    cfg['E1'] = '工艺强度等级'
    cfg['G1'] = '税率'
    for c in ('A1', 'B1', 'D1', 'E1', 'G1'):
        cfg[c].font = _FONT_BOLD

    # 8 档岩石基准含税价
    for i, (rock, price) in enumerate(RECOMMEND_BASELINE_INCL.items(), start=2):
        cfg.cell(row=i, column=1, value=rock)
        cfg.cell(row=i, column=2, value=float(price))

    # 工艺强度等级
    for i, (proc, lv) in enumerate(PROCESS_LEVEL.items(), start=2):
        cfg.cell(row=i, column=4, value=proc)
        cfg.cell(row=i, column=5, value=int(lv))

    # 税率（从主参数库读，便于联动）
    cfg['G2'] = "='基础核心-参数库'!$B$20"

    # 列宽
    cfg.column_dimensions['A'].width = 16
    cfg.column_dimensions['B'].width = 28
    cfg.column_dimensions['D'].width = 22
    cfg.column_dimensions['E'].width = 14
    cfg.column_dimensions['G'].width = 10

    # 顶部使用说明
    cfg['A12'] = '说明：本 sheet 由 model_core.py 自动生成，请勿手动修改。'
    cfg['A13'] = '主表 B65「综合单价(含税)」公式会从本表读取基准价和工艺等级，'
    cfg['A14'] = '在用户改了岩石级别/工艺类型后自动重算"阶梯定价校正"。'
    for r in (12, 13, 14):
        cfg.cell(row=r, column=1).alignment = _ALIGN_WRAP

    # ② 改造主表 B65 为校正后活公式（其他单元格保持 V8 原算公式不动）
    if main_sheet_name not in wb.sheetnames:
        return
    ws = wb[main_sheet_name]

    cfg_q = f"'{cfg_sheet_name}'"
    # 公式片段
    f_v8_excl = "SUM(B57:B63)"
    f_v8_incl = f"{f_v8_excl}*(1+'基础核心-参数库'!$B$20)"
    f_user_lv = f"IFERROR(VLOOKUP(B12,{cfg_q}!$D$2:$E$6,2,0),0)"
    f_rec_lv  = f"IFERROR(VLOOKUP(B13,{cfg_q}!$D$2:$E$6,2,0),0)"
    f_gap     = f"MAX(0,({f_rec_lv})-({f_user_lv}))"
    f_factor  = f"IF({f_gap}>0,POWER(2,{f_gap}),1)"
    f_base    = f"IFERROR(VLOOKUP(B4,{cfg_q}!$A$2:$B$9,2,0),0)"
    # v10 修复（病灶1）：强制走阶梯定价 baseline × 2^gap
    # 旧公式：f_correct = MAX(基准×系数, V8原算) → 硬岩+松土时V8原算更高被走else
    # 新公式：f_correct = 基准×系数（直接强制），V8原算保留在 D70 显示作为对照
    f_correct = f"{f_base}*{f_factor}"

    # B64：综合单价(不含税) → 校正后/(1+税率)（错配时）或 V8 原算之和（无错配时）
    ws['B64'] = f"=IF({f_gap}>0,{f_correct}/(1+'基础核心-参数库'!$B$20),{f_v8_excl})"
    # B65：综合单价(含税) → 校正后含税（错配时）或 V8 原算含税（无错配时）
    ws['B65'] = f"=IF({f_gap}>0,{f_correct},{f_v8_incl})"

    # 标记 B64/B65 是校正后数字（视觉提示）
    ws['B64'].fill = _FILL_OK
    ws['B65'].fill = _FILL_OK
    ws['B65'].font = Font(bold=True, color='C00000', size=12)

    # ③ 主表 D-G 列追加"校正过程说明"区，让用户能看到计算依据
    # 不动 A-C 已有 V8 行布局，把说明区放到 D64:G70
    ws['D64'] = '校正过程（改岩石/工艺自动重算）'
    ws['D64'].font = _FONT_TITLE
    ws.merge_cells('D64:G64')

    rows = [
        ('推荐基准含税价(元/方)', f"={f_base}"),
        ('您工艺强度等级',        f"={f_user_lv}"),
        ('推荐工艺强度等级',      f"={f_rec_lv}"),
        ('档位差(偏软档数)',      f"={f_gap}"),
        ('阶梯系数(2^档位差)',    f"={f_factor}"),
        ('V8 原算含税(元/方)',    f"={f_v8_incl}"),
    ]
    for i, (label, formula) in enumerate(rows, start=65):
        ws.cell(row=i, column=4, value=label).font = _FONT_BOLD
        ws.cell(row=i, column=4).alignment = _ALIGN_WRAP
        ws.cell(row=i, column=4).fill = _FILL_NOTE
        ws.cell(row=i, column=5, value=formula)
        ws.cell(row=i, column=5).alignment = _ALIGN_WRAP
        ws.cell(row=i, column=5).fill = _FILL_NOTE
        ws.merge_cells(start_row=i, end_row=i, start_column=5, end_column=7)

    # 列宽（仅当宽度未设过时设置）
    try:
        ws.column_dimensions['D'].width = max(ws.column_dimensions['D'].width or 0, 26)
        ws.column_dimensions['E'].width = max(ws.column_dimensions['E'].width or 0, 16)
    except Exception:
        ws.column_dimensions['D'].width = 26
        ws.column_dimensions['E'].width = 16


def _write_changelog_sheet(wb) -> None:
    """写入「📜 更新日志-CHANGELOG」sheet，把 CHANGELOG_V10 逐条沉淀到 Excel。

    每次模型升级追加新版本 CHANGELOG_VXX 都会自动追加到这里，便于用户追溯升级历史。
    """
    sheet_name = '📜 更新日志-CHANGELOG'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # 标题行
    ws['A1'] = '📜 v10 修复版 — 更新日志'
    ws['A1'].font = Font(bold=True, size=14, color='305496')
    ws.merge_cells('A1:G1')

    ws['A2'] = '说明：本表逐条记录 v10 相对 v9-方案K 的修复内容，含修改前/后效果对比与文件位置，便于版本追溯。'
    ws['A2'].alignment = _ALIGN_WRAP
    ws['A2'].font = Font(italic=True, color='666666', size=10)
    ws.merge_cells('A2:G2')

    # 表头
    headers = ['#', '日期', '严重程度', '问题/改动', '修改前', '修改后', '改动文件']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = _FONT_HEAD
        c.fill = _FILL_HEAD
        c.alignment = _ALIGN_CTR

    # 逐条写入
    severity_fill = {
        '🔴 严重':   PatternFill('solid', fgColor='F8CBAD'),
        '🟡 中等':   PatternFill('solid', fgColor='FFF2CC'),
        '🟠 反直觉': PatternFill('solid', fgColor='FCE4D6'),
        '🟢 数据':   PatternFill('solid', fgColor='E2EFDA'),
        '⚪ 工具':   PatternFill('solid', fgColor='F2F2F2'),
        '⚪ 体验':   PatternFill('solid', fgColor='D9E1F2'),
    }
    for i, item in enumerate(CHANGELOG_V10, start=5):
        row_fill = severity_fill.get(item['severity'], _FILL_NOTE)
        ws.cell(row=i, column=1, value=item['no'])
        ws.cell(row=i, column=2, value=item['date'])
        ws.cell(row=i, column=3, value=item['severity'])
        ws.cell(row=i, column=4, value=f"{item['title']}\n— {item['detail']}")
        ws.cell(row=i, column=5, value=item['before'])
        ws.cell(row=i, column=6, value=item['after'])
        ws.cell(row=i, column=7, value=item['files'])
        for col in range(1, 8):
            cell = ws.cell(row=i, column=col)
            cell.alignment = _ALIGN_WRAP
            cell.fill = row_fill

    # 列宽
    widths = {'A': 5, 'B': 12, 'C': 11, 'D': 50, 'E': 36, 'F': 36, 'G': 38}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 行高（让长文本能完全显示）
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[4].height = 22
    for i in range(5, 5 + len(CHANGELOG_V10)):
        ws.row_dimensions[i].height = 72

    # 底部签名
    foot_row = 5 + len(CHANGELOG_V10) + 1
    ws.cell(row=foot_row, column=1, value='— 生成自 model_core.py / 修复版本 v10 / 2026-06-03 —')
    ws.cell(row=foot_row, column=1).font = Font(italic=True, color='999999', size=9)
    ws.merge_cells(start_row=foot_row, end_row=foot_row, start_column=1, end_column=7)


def _hide_orphan_params(wb) -> None:
    """扫描「基础核心-参数库」sheet，将「未被任何公式引用」的参数 cell 灰显标记。

    工程意义：参数库里大量历史预留/废弃/备用参数会让用户误以为"改了就有效"，
    但这些 cell 实际上从来不进任何公式。本函数把它们灰显并加 sheet 级注释，
    既不删数据（保留扩展空间），又给用户清晰的视觉提示。

    扫描算法：
        ① 遍历所有 sheet 公式，提取所有 cell 引用（含跨表区域 + 本表区域）
        ② 取出对参数库的引用集合
        ③ 参数库内所有非空数值 cell，若不在引用集合 → 标灰
    """
    import re
    from openpyxl.utils import column_index_from_string, get_column_letter

    param_sheet_name = '基础核心-参数库'
    if param_sheet_name not in wb.sheetnames:
        return

    def expand_refs(formula, src_sheet):
        refs = set()
        # ① 跨表区域 'xx'!$A$1:$B$10
        rng_cross = re.compile(
            r"'?([^'!]*?)'?\s*!\s*\$?([A-Z]+)\$?(\d+)\s*:\s*\$?([A-Z]+)\$?(\d+)",
            re.IGNORECASE,
        )
        for m in rng_cross.finditer(formula):
            sname, c1, r1, c2, r2 = m.groups()
            c1i, c2i = column_index_from_string(c1), column_index_from_string(c2)
            for col in range(min(c1i, c2i), max(c1i, c2i) + 1):
                for row in range(min(int(r1), int(r2)), max(int(r1), int(r2)) + 1):
                    refs.add((sname, f"{get_column_letter(col)}{row}"))
        f1 = rng_cross.sub('', formula)
        # ② 跨表单 cell 'xx'!$A$1
        cross_single = re.compile(
            r"'?([^'!,\s\(\)\+\-\*/=<>%&]+?)'?\s*!\s*\$?([A-Z]+)\$?(\d+)",
            re.IGNORECASE,
        )
        for m in cross_single.finditer(f1):
            sname, col, row = m.groups()
            refs.add((sname, f"{col.upper()}{row}"))
        f2 = cross_single.sub('', f1)
        # ③ 本表区域 A1:B10
        rng_local = re.compile(r"\$?([A-Z]+)\$?(\d+)\s*:\s*\$?([A-Z]+)\$?(\d+)")
        for m in rng_local.finditer(f2):
            c1, r1, c2, r2 = m.groups()
            c1i, c2i = column_index_from_string(c1), column_index_from_string(c2)
            for col in range(min(c1i, c2i), max(c1i, c2i) + 1):
                for row in range(min(int(r1), int(r2)), max(int(r1), int(r2)) + 1):
                    refs.add((src_sheet, f"{get_column_letter(col)}{row}"))
        f3 = rng_local.sub('', f2)
        # ④ 本表单 cell
        single = re.compile(r"(?<![A-Z!])\$?([A-Z]+)\$?(\d+)")
        for m in single.finditer(f3):
            col, row = m.groups()
            refs.add((src_sheet, f"{col.upper()}{row}"))
        return refs

    # 全表扫描所有 cell 引用
    all_refs: set = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith('='):
                    all_refs.update(expand_refs(v[1:], ws.title))

    # 参数库被引用 cell（含别名）
    param_aliases = {param_sheet_name, '参数库'}
    param_referenced = {coord for sname, coord in all_refs if sname in param_aliases}

    # 参数库所有非空数值 cell（不含公式）
    ws_p = wb[param_sheet_name]
    orphan_fill = PatternFill('solid', fgColor='D9D9D9')
    orphan_font = Font(color='999999', italic=True, size=10)

    orphan_count = 0
    for row in ws_p.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and v.startswith('='):
                continue  # 公式不动
            if not isinstance(v, (int, float)):
                continue  # 字符串标签不动
            if cell.coordinate in param_referenced:
                continue  # 已被引用
            # 孤儿：灰显
            cell.fill = orphan_fill
            cell.font = orphan_font
            orphan_count += 1

    # 在参数库顶部加汇总说明（如果有孤儿）
    if orphan_count > 0:
        # 找第 1 个空行的下面位置（用 AC1 避免冲突）
        ws_p['AC1'] = f'⚪ 灰显说明'
        ws_p['AC1'].font = Font(bold=True, color='999999', size=10)
        ws_p['AC2'] = f'共 {orphan_count} 个参数 cell 当前未被任何公式引用'
        ws_p['AC2'].font = Font(color='999999', size=9)
        ws_p['AC3'] = '已用浅灰显示——改这些值不会影响计算结果'
        ws_p['AC3'].font = Font(color='999999', size=9)
        ws_p['AC4'] = '保留它们是为后续扩展（如新工艺/新设备）预留空间'
        ws_p['AC4'].font = Font(color='999999', size=9)
        ws_p.column_dimensions['AC'].width = 42


def export_xlsx(params: Dict[str, Any] = None) -> bytes:
    """
    给一组用户参数，返回完整 xlsx 字节流（**交互式版本**）。

    交互特性：用户拿到 Excel 后，改岩石/工艺/钻孔/运距等任何参数，
    Excel 会自动重算 B65「综合单价(含税)」，**含校正层（阶梯定价）也会自动跟着算**。

    单一数据源：RECOMMEND_BASELINE_INCL / PROCESS_LEVEL 来自 model_core.py，
    生成 Excel 时写进隐藏配置 sheet「⚙️校正配置」，所以 model_core 改了基准价/工艺等级，
    用户下载新版 Excel 后改参数算出来的结果也会跟着 model_core 走。

    具体步骤：
        1. _build_and_calc 跑 libreoffice 重算（拿到 V8 原算公式缓存）
        2. 读出 OUTPUT_CELLS 用于构造审计 sheet（这部分仍是"下载时快照"）
        3. _inject_correction_layer 注入校正活公式 + 隐藏配置 sheet（核心交互能力）
        4. 追加「📋 工艺设备匹配度校验」「💰 校正后定价说明」两个 sheet（快照）
        5. BytesIO 保存返回字节流
    """
    params = params or {}
    with tempfile.TemporaryDirectory() as workdir:
        recalc_path = _build_and_calc(params, workdir)

        # ① 读出 V8 原算值（用于审计 sheet 的"下载时快照"）
        wb_ro = load_workbook(recalc_path, data_only=True)
        out: Dict[str, Any] = {}
        for key, (sheet, cell) in OUTPUT_CELLS.items():
            try:
                out[key] = wb_ro[sheet][cell].value
            except Exception:
                out[key] = None
        out['inputs'] = {**DEFAULT_PARAMS, **params}
        wb_ro.close()

        # ② 校正 + 审计（与网页同源，用于审计 sheet 快照）
        apply_penalty(out)
        audit = build_audit_report(out)

        # ③ 重新打开（保留公式），注入校正层活公式 + 写审计 sheet
        wb = load_workbook(recalc_path, data_only=False)

        # 核心：把校正逻辑做成活公式 → Excel 改参数能自动重算
        _inject_correction_layer(wb)

        # 审计 sheet 仍是"下载时快照"
        _write_audit_sheet(wb, audit, out)
        _write_correction_sheet(wb, out, audit)

        # v10 新增：CHANGELOG 永久沉淀 + 孤儿参数灰显（必须在所有 sheet 写完后再扫描）
        _write_changelog_sheet(wb)
        try:
            _hide_orphan_params(wb)
        except Exception as e:
            # 孤儿扫描出错不影响主流程
            print(f'[export_xlsx] 孤儿参数扫描跳过: {e}', file=sys.stderr)

        # ④ 保存到字节流
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        return buf.getvalue()


# ---------- 自检 ----------
if __name__ == '__main__':
    print("[1/3] 生成模板...")
    _ensure_template()
    print(f"      模板路径: {_TEMPLATE_PATH}  ({os.path.getsize(_TEMPLATE_PATH)} bytes)")

    print("[2/3] 默认工况（Ⅰ极软岩 + 直接开挖）...")
    r = compute()
    print(f"      含税综合单价: {r['price_incl_tax']:.4f} 元/方  （预期 ≈ 8.92）")
    print(f"      不含税:       {r['price_excl_tax']:.4f}")
    print(f"      挖装费:       {r['cost_excavate']:.4f}")
    print(f"      运输费:       {r['cost_transport']:.4f}")

    print("[3/3] 黑山合同对标工况（Ⅲ较软岩 + 直接开挖）...")
    r2 = compute({'rock_level': 'Ⅲ较软岩'})
    print(f"      含税综合单价: {r2['price_incl_tax']:.4f} 元/方  （预期 ≈ 11.88）")

    print("\n✅ 自检通过")
